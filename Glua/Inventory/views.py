import csv
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from django.db.models import Sum, F, Q
from .models import Drug, Sale, Stocked, LockedProduct, MarketingItem, IssuedItem, PickingList, Cannister, IssuedCannister, Client
from .forms import DrugCreation
from django.contrib import messages
from django.views.generic import ListView, UpdateView
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from django.utils.timezone import now, localtime, make_aware
from datetime import datetime, timedelta, time
from django.db.models import Count
from django.http import JsonResponse
from django.contrib.auth.models import User
from django.contrib.sessions.models import Session
from django.contrib.auth import logout
from django.core.paginator import Paginator
from django.utils.dateparse import parse_date
from django.views.decorators.csrf import csrf_exempt
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from django.conf import settings
import os


# Create your views here.

@login_required
def home(request):
    today = timezone.now().date()

    # Get the products expiring within the next 10 days
    expiring_soon = Drug.objects.filter(expiry_date__lte=today + timedelta(days=10), expiry_date__gt=today)

    # Get the products with stock below the reorder level
    low_stock = Drug.objects.filter(stock__lte=F('reorder_level'))

    # Pagination handling
    per_page = request.GET.get('per_page', 10)  # Default to 10 per page
    page_number = request.GET.get('page', 1)    # Get the current page number
    drugs = Drug.objects.all().order_by('name')# Get all drugs ordered by name

    # Create a paginator and get the page object for the current page
    paginator = Paginator(drugs, per_page)
    page_obj = paginator.get_page(page_number)

    # Get all clients for the dropdown
    clients = Client.objects.all().order_by('name')

    # Check if the modal has already been shown in this session
    show_modal = not request.session.get('modal_shown', False)  # Only show modal if 'modal_shown' is not set or False

    if show_modal:
        print("Modal will be shown")
        request.session['modal_shown'] = True  # Set the session variable to True after showing the modal
        request.session.modified = True  # Ensure the session is saved
    else:
        print("Modal already shown in this session")

    # Pass these to the template
    context = {
        'drugs': page_obj,  # Pass the paginated drugs
        'clients': clients,  # Pass clients for dropdown
        'expiring_soon': expiring_soon,
        'low_stock': low_stock,
        'show_modal': show_modal,  # Pass this flag to the template
    }

    return render(request, 'Inventory/home.html', context)



@login_required
def createDrug(request):
    try:
        if request.method == 'POST':
            form = DrugCreation(request.POST)
            if form.is_valid():
                name = form.cleaned_data.get('name')
                form.save()
                messages.success(
                    request, f'{name} has been successfully added to the inventory'
                )
                return redirect('home')  # Redirect to the homepage or another page
        else:
            form = DrugCreation()

        context = {'form': form}
        return render(request, 'Inventory/create.html', context)

    except Exception as e:
        messages.error(request, f'An error occurred: {str(e)}')
        return redirect('home')  # Redirect to 'home' in case of an error


@login_required
def addStock(request, pk):
    drug = Drug.objects.get(id=pk)
    supp = request.POST.get('supplier')
    amount_added = int(request.POST.get('added'))
    drug.stock += amount_added
    Stocked.objects.create(
        drug_name=drug, supplier=supp, staff=request.user, number_added=amount_added, total=drug.stock)
    drug.save()
    messages.success(request, f'{amount_added} {drug.name} added')
    return redirect('stocking')


class stockingListView(ListView):
    model = Drug
    context_object_name = 'drugs'
    paginate_by = 5
    template_name = 'Inventory/stock.html'
    ordering = ['name']


@login_required
def sellDrug(request, pk):
    if request.method == 'POST':
        quantity = float(request.POST.get('quantity'))
        client_id = request.POST.get('client')
        drug = get_object_or_404(Drug, pk=pk)

        if not client_id:
            messages.error(request, 'Please select a client')
            return redirect('home')

        try:
            client = Client.objects.get(id=client_id)
        except Client.DoesNotExist:
            messages.error(request, 'Selected client does not exist')
            return redirect('home')

        if drug.stock >= quantity:
            # Update drug stock
            drug.stock -= quantity
            drug.save()

            # Create sale record
            Sale.objects.create(
                seller=request.user,
                drug_sold=drug.name,
                client=client,
                batch_no=drug.batch_no,
                quantity=quantity,
                remaining_quantity=drug.stock
            )

            messages.success(request, f'{quantity} {drug.name} sold to {client.name}')
        else:
            messages.error(request, 'Not enough stock available')

        return redirect('home')

@login_required
def lockDrug(request, pk):
    if request.method == 'POST':
        quantity = float(request.POST.get('quantity'))
        client_id = request.POST.get('client')
        drug = get_object_or_404(Drug, pk=pk)

        if not client_id:
            messages.error(request, 'Please select a client')
            return redirect('home')

        try:
            client = Client.objects.get(id=client_id)
        except Client.DoesNotExist:
            messages.error(request, 'Selected client does not exist')
            return redirect('home')

        if drug.stock >= quantity:
            # Lock product by creating a LockedProduct record
            LockedProduct.objects.create(
                drug=drug,
                locked_by=request.user,
                quantity=quantity,
                client=client
            )

            drug.stock -= quantity
            drug.save()

    #         last_sale = Sale.objects.filter(drug_sold=drug).order_by('-date_sold').first()

    # # If a sale exists, add the locked quantity to the remaining_quantity
    #         if last_sale:
    #             if last_sale.remaining_quantity is None:
    #                 last_sale.remaining_quantity = 0  # Ensure it's initialized
    #             last_sale.remaining_quantity = drug.stock
    #             last_sale.save()

            # Reduce stock

            messages.success(request, f'{quantity} {drug.name} locked.')
        else:
            messages.error(request, 'Not enough stock to lock')

        return redirect('home')


def search(request):
    drugs = Drug.objects.all().order_by('name')
    query = request.POST.get('q')

    if query:
        drugs = Drug.objects.filter(
            Q(name__icontains=query) | Q(batch_no__icontains=query))

    context = {'drugs': drugs}
    return render(request, 'Inventory/home.html', context)


def binsearch(request):
    bins = Sale.objects.all().order_by('drug_sold')
    
    # Get search query from GET request or fallback to POST request
    query = request.GET.get('search') or request.POST.get('quiz')

    print("Search Query:", query)  # Debugging the query received

    if query:
        bins = bins.filter(
            Q(drug_sold__icontains=query) |
            Q(batch_no__icontains=query) |
            Q(client__name__icontains=query)
        ).order_by('date_sold')
    return render(request, 'Inventory/bin.html', {'sales': bins})



def searchstock(request):
    drugs = Drug.objects.all().order_by('name')
    query = request.POST.get('s')

    if query:
        drugs = Drug.objects.filter(Q(name__icontains=query)).order_by('name')

    context = {'drugs': drugs}
    return render(request, 'Inventory/stock.html', context)


def salehistory(request):
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    if start_date and end_date:
        sales = Sale.objects.filter(
            date_sold__range=[start_date, end_date]).order_by('-date_sold')
        if sales:
            total_sales = sales.aggregate(
                total=Sum(F('quantity') * F('sale_price')))['total']
            bought_at = sales.aggregate(
                total=Sum(F('quantity') * F('buying_price')))['total']
            profit = total_sales - bought_at
            context = {'sales': sales, 'profit': profit, 'total_sales': total_sales}
        else:
            messages.success(
                request, 'Sorry no sales were done within those dates')
            context = {}
            return redirect('history')
    else:
        context = {}
    return render(request, 'Inventory/history.html', context)


def todaysales(request):
    today = datetime.now().date()
    tomorrow = today + timedelta(1)
    start_date = datetime.combine(today, time())
    end_date = datetime.combine(tomorrow, time())

    if start_date and end_date:
        sales = Sale.objects.filter(
            date_sold__range=[start_date, end_date]).order_by('-date_sold')
        if sales:
            total_sales = sales.aggregate(
                total=Sum(F('quantity') * F('sale_price')))['total']
            bought_at = sales.aggregate(
                total=Sum(F('quantity') * F('buying_price')))['total']
            profit = total_sales - bought_at
            context = {'sales': sales, 'profit': profit, 'total_sales': total_sales}
        else:
            messages.success(request, 'Sorry no sales were done today')
            context = {}
            return redirect('history')
    else:
        context = {}
    return render(request, 'Inventory/today.html', context)


def StockAdded(request):
    start_date = request.GET.get('date_start')
    end_date = request.GET.get('date_end')
    if start_date and end_date:
        glua_stocked_days = Stocked.objects.filter(
            date_added__range=[start_date, end_date]).order_by('-date_added')
        context = {'stocked': glua_stocked_days}
    else:
        context = {}

    return render(request, 'Inventory/stocked.html', context)


class modifyDrugUpdateView(UpdateView):
    template_name = 'Inventory/create.html'
    model = Drug
    fields = ['name', 'stock', 'batch_no']
    success_url = "/"


def bin_report(request):
    # Get all sales ordered by date sold
    sales = Sale.objects.all().order_by('-date_sold')
    
    # Get date range filters from the request
    start_date = request.POST.get('start_date')
    end_date = request.POST.get('end_date')
    
    # Filter sales based on the date range, if provided
    if start_date and end_date:
        try:
            start_date = datetime.strptime(start_date, "%Y-%m-%d").date()
            end_date = datetime.strptime(end_date, "%Y-%m-%d").date()
            sales = sales.filter(date_sold__range=(start_date, end_date)).order_by('date_sold')
        except ValueError:
            pass  # Ignore invalid dates

    # Pagination setup
    per_page = int(request.GET.get('per_page', 10))  # Default to 10 items per page
    paginator = Paginator(sales, per_page)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)

    return render(request, 'Inventory/bin.html', {'sales': page_obj})

@login_required
def download_bin_report_excel(request):
    """Export bin report as styled Excel file using template"""
    try:
        # Load template
        template_path = os.path.join(settings.BASE_DIR, 'Inventory transfer record template11.xlsx')
        wb = load_workbook(template_path)
        ws = wb['Template']  # Work with Template sheet
        
        # Delete rows 3 and 4 (template placeholder data)
        ws.delete_rows(3, 2)
        
        # Get all sales
        sales = Sale.objects.all().order_by('-date_sold')
        
        # Apply search filter if provided
        search_query = request.GET.get('search') or request.GET.get('q') or request.POST.get('q')
        if search_query:
            sales = sales.filter(
                Q(drug_sold__icontains=search_query) |
                Q(batch_no__icontains=search_query) |
                Q(client__name__icontains=search_query)
            )
        
        # Apply date range filters
        start_date = request.GET.get('start_date') or request.POST.get('start_date')
        end_date = request.GET.get('end_date') or request.POST.get('end_date')
        
        if start_date and end_date:
            try:
                start_date_obj = datetime.strptime(start_date, "%Y-%m-%d").date()
                end_date_obj = datetime.strptime(end_date, "%Y-%m-%d").date()
                sales = sales.filter(date_sold__range=(start_date_obj, end_date_obj)).order_by('date_sold')
            except ValueError:
                pass
        else:
            sales = sales.order_by('-date_sold')
        
        # Start writing data from row 3 (after headers)
        row_num = 3
        for sale in sales:
            ws.cell(row=row_num, column=1, value=sale.drug_sold if sale.drug_sold else '')
            ws.cell(row=row_num, column=2, value=sale.batch_no if sale.batch_no else '')
            ws.cell(row=row_num, column=3, value=sale.quantity)
            ws.cell(row=row_num, column=4, value=sale.client.name if sale.client else '')
            ws.cell(row=row_num, column=5, value=sale.date_sold.strftime('%Y-%m-%d') if sale.date_sold else '')
            row_num += 1
        
        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="bin_report.xlsx"'
        wb.save(response)
        return response
        
    except Exception as e:
        return HttpResponse(f"Error generating Excel file: {str(e)}", status=500)


@login_required
def download_bin_card_excel(request):
    """Export bin card as styled Excel file using template"""
    try:
        # Load template
        template_path = os.path.join(settings.BASE_DIR, 'Inventory transfer record template11.xlsx')
        wb = load_workbook(template_path)
        ws = wb['Template']  # Work with Template sheet
        
        # Delete rows 3 and 4 (template placeholder data)
        ws.delete_rows(3, 2)
        
        # Get all issued cannisters
        issued_cannisters = IssuedCannister.objects.all().order_by('-date_issued')
        
        # Apply search filter if provided
        search_query = request.GET.get('search') or request.POST.get('search')
        if search_query:
            issued_cannisters = issued_cannisters.filter(
                Q(name__icontains=search_query) |
                Q(batch_no__icontains=search_query) |
                Q(staff_on_duty__username__icontains=search_query)
            )
        
        # Apply date range filters
        start_date = request.GET.get('start_date') or request.POST.get('start_date')
        end_date = request.GET.get('end_date') or request.POST.get('end_date')
        
        if start_date and end_date:
            try:
                start_date_obj = datetime.strptime(start_date, "%Y-%m-%d").date()
                end_date_obj = datetime.strptime(end_date, "%Y-%m-%d").date()
                issued_cannisters = issued_cannisters.filter(
                    date_issued__range=(start_date_obj, end_date_obj)
                ).order_by('date_issued')
            except ValueError:
                pass
        else:
            issued_cannisters = issued_cannisters.order_by('-date_issued')
        
        # Start writing data from row 3 (after headers)
        row_num = 3
        for cannister in issued_cannisters:
            ws.cell(row=row_num, column=1, value=cannister.name if cannister.name else '')
            ws.cell(row=row_num, column=2, value=cannister.batch_no if cannister.batch_no else '')
            ws.cell(row=row_num, column=3, value=cannister.staff_on_duty.username if cannister.staff_on_duty else '')
            ws.cell(row=row_num, column=4, value=cannister.client.name if cannister.client else '')
            ws.cell(row=row_num, column=5, value=cannister.quantity)
            ws.cell(row=row_num, column=6, value=cannister.balance)
            ws.cell(row=row_num, column=7, value=cannister.date_issued.strftime('%Y-%m-%d') if cannister.date_issued else '')
            ws.cell(row=row_num, column=8, value=cannister.date_returned.strftime('%Y-%m-%d') if cannister.date_returned else '')
            row_num += 1
        
        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="bin_card.xlsx"'
        wb.save(response)
        return response
        
    except Exception as e:
        return HttpResponse(f"Error generating Excel file: {str(e)}", status=500)


@login_required
def dashboard(request):
    today = timezone.now().date()

    # Get the expired products (expiry date is in the past)
    expired_drugs = Drug.objects.filter(expiry_date__lt=today, stock__gt=0)

    # Get the products expiring within the next 10 days
    expiring_soon = Drug.objects.filter(expiry_date__lte=today + timedelta(days=180), expiry_date__gt=today, stock__gt=0).order_by('expiry_date')

    # Get the products with stock below the reorder level
    low_stock = Drug.objects.filter(stock__lte=F('reorder_level'), stock__gt=0)
    out_of_stock = Drug.objects.filter(stock=0)

    # Check if the modal should be shown (only when there are low stock or expiring soon products)
    show_modal = False
    if low_stock.exists() or expiring_soon.exists() or out_of_stock.exists():
        show_modal = not request.session.get('modal_shown', False)  # Only show modal if 'modal_shown' is not set or False

    if show_modal:
        request.session['modal_shown'] = True  # Set the session variable to True after showing the modal
        request.session.modified = True  # Ensure the session is saved

    # Summary Data
    total_products = Drug.objects.count()
    low_stock_products = Drug.objects.filter(stock__lte=F('reorder_level'), stock__gt=0).count()
    out_of_stock_products = Drug.objects.filter(stock=0).count()
    zero_stock_products = Drug.objects.filter(stock__lte=5).count()
    locked_products = LockedProduct.objects.all().count()
    marketing_items = MarketingItem.objects.all().count()
    total_picking_list = PickingList.objects.all().count()
    cannisters = Cannister.objects.all().count()

    # Top Sold Products
    top_sold_products = (
        Sale.objects.values("drug_sold")
        .annotate(total_quantity=Sum("quantity"))
        .order_by("-total_quantity")[:210000]
    )

    # Calculate the total count of expired and expiring soon drugs
    total_expiring_count = expired_drugs.count() + expiring_soon.count()

    context = {
        'total_products': total_products,
        'low_stock_products': low_stock_products,
        'out_of_stock_products': out_of_stock_products,
        'zero_stock_products': zero_stock_products,
        'top_sold_products': top_sold_products,
        'expired_drugs_count': expired_drugs.count(),  # Add the count of expired drugs
        'expiring_soon_count': expiring_soon.count(),  # Add the count of expiring soon drugs
        'total_expiring_count': total_expiring_count,  # Add the total count of expired and expiring soon drugs
        'expired_drugs': expired_drugs,  # Pass expired drugs to the template
        'expiring_soon': expiring_soon,  # Pass expiring soon drugs to the template
        'low_stock': low_stock,
        'show_modal': show_modal,
        'locked_products': locked_products,
        'marketing_items': marketing_items,
        'total_picking_list': total_picking_list,
        'cannisters':cannisters,
        'out_of_stock':out_of_stock
    }
    print("Context out_of_stock:", list(out_of_stock))
    return render(request, 'Inventory/dashboard.html', context)


@login_required
def low_stock_view(request):
    # Get the products with stock below or equal to the reorder level
    low_stock = Drug.objects.filter(stock__lte=F('reorder_level'), stock__gt=0)

    context = {
        'low_stock': low_stock
    }

    return render(request, 'Inventory/lowstock.html', context)

def get_online_offline_users(request):
    # Get all users
    all_users = User.objects.all()

    # Split them into online and offline users
    online_users = [user.username for user in all_users if user.is_active]  # Assuming 'is_active' indicates online status
    offline_users = [user.username for user in all_users if not user.is_active]  # Assuming 'is_active' indicates offline status

    # Return as JSON response
    return JsonResponse({
        'online_users': online_users,
        'offline_users': offline_users,
    })

@login_required
def locked_products(request):
    """
    Display the list of locked products in ascending order by the drug name.
    """
    # Fetch all locked products and order by the drug's name
    locked_products = LockedProduct.objects.all().order_by('-date_locked')
    return render(request, 'Inventory/locked.html', {'locked_products': locked_products})

@login_required
def post_locked_product(request, lock_id):
    """
    Handle the action of posting a locked product.
    """
    lock = get_object_or_404(LockedProduct, id=lock_id)
    if request.method == 'POST':
        quantity = lock.quantity
        client = lock.client  # Assuming 'locked_by' is a User and you need their username as the client.
        drug = lock.drug

        # Create sale record
        Sale.objects.create(
            seller=request.user,
            drug_sold=drug.name,
            client=client,
            batch_no=drug.batch_no,
            quantity=quantity,
            remaining_quantity=drug.stock
        )

        # Delete the locked product
        lock.delete()

        # Display a success message
        messages.success(request, f'{quantity} {drug.name} sold to {client} and lock removed.')

        # Redirect to the locked products page
        return redirect('locked_products')


@login_required
def unlock_product(request, lock_id):
    """
    Handle the action of unlocking a product.
    """
    # Fetch the locked product instance
    lock = get_object_or_404(LockedProduct, id=lock_id)

    # Add the locked quantity back to the drug's stock
    drug = lock.drug
    if lock.quantity:  # Ensure the quantity is not None or empty
        drug.stock += int(lock.quantity)  # Adding the locked quantity back to the stock
        drug.save()  # Save the updated stock to the database

    # Fetch the last sale entry for this drug
    # last_sale = Sale.objects.filter(drug_sold=drug).order_by('-date_sold').first()

    # # If a sale exists, add the locked quantity to the remaining_quantity
    # if last_sale:
    #     if last_sale.remaining_quantity is None:
    #         last_sale.remaining_quantity = 0  # Ensure it's initialized
    #     last_sale.remaining_quantity = drug.stock
    #     last_sale.save()  # Save the updated remaining quantity

    # Delete the locked product after updating the stock and sales
    lock.delete()

    # Redirect to the locked products page
    messages.success(request, f"{lock.quantity} {drug.name} unlocked and added back to stock.")
    return redirect('locked_products')


@login_required
def locked_search(request):
    query = request.POST.get('quiz', '')  # Retrieve the search query from the form
    locked_products = LockedProduct.objects.filter(
        Q(drug__name__icontains=query) | Q(locked_by__username__icontains=query)
    ).order_by('-date_locked')  # Search for drug name or locked_by username containing the query (case-insensitive)

    return render(request, 'Inventory/locked.html', {'locked_products': locked_products})

@login_required
def user_management(request):
    """
    Display the user management page with a form to add new users
    and a list of online/offline users, with their login and logout times.
    Automatically log out users after 1 minute of inactivity.
    """
    # Get current time (timezone-aware)
    # current_time = now()

    # # Convert last_activity to timezone-aware datetime
    # last_activity = request.session.get('last_activity', None)
    # if last_activity:
    #     # Convert the last_activity from string to datetime
    #     last_activity = datetime.strptime(last_activity, '%Y-%m-%d %H:%M:%S')
    #     # Make it timezone-aware
    #     last_activity = make_aware(last_activity)

    # # Check if the user has been inactive for more than 1 minute
    # inactivity_duration = (current_time - last_activity).total_seconds() if last_activity else 0

    # if inactivity_duration > 60:
    #     # If more than 1 minute has passed, log out the user
    #     logout(request)  # This will log out the user
    #     return render(request, 'Inventory/user_management.html', {'message': 'You have been logged out due to inactivity.'})

    # # Update last activity timestamp as a string
    # request.session['last_activity'] = current_time.strftime('%Y-%m-%d %H:%M:%S')

    # Get all users
    users = User.objects.all()

    # Get session information to track online users and their login times
    sessions = Session.objects.filter(expire_date__gte=now())
    online_user_ids = [session.get_decoded().get('_auth_user_id') for session in sessions]

    # Annotate users with their online/offline status, login time, and logout time
    for user in users:
        user.is_online = str(user.id) in online_user_ids
        if user.is_online:
            # If the user is online, show login time
            user.login_time = localtime(user.last_login).strftime('%Y-%m-%d %H:%M:%S') if user.last_login else None
            user.logout_time = None  # No logout time for online users
        else:
            # If the user is offline, show logout time
            user.login_time = None  # No login time for offline users
            user.logout_time = localtime(user.last_login).strftime('%Y-%m-%d %H:%M:%S') if user.last_login else "N/A"

    context = {
        'users': users,
    }

    return render(request, 'Inventory/user_management.html', context)

@login_required
def add_user(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        email = request.POST.get('email')
        password = request.POST.get('password')

        # Validate inputs
        if not username or not email or not password:
            messages.error(request, "All fields are required.")
            return render(request, 'add_user.html')  # Ensure form is empty if error

        if User.objects.filter(Q(username=username) | Q(email=email)).exists():
            messages.error(request, "A user with this username or email already exists.")
            return render(request, 'add_user.html')  # Ensure form is empty if error

        User.objects.create_user(username=username, email=email, password=password)
        messages.success(request, "New user added successfully.")
        return redirect('user_management')  # Redirect to a success page

    # If the request method is not POST, render the page with an empty form
    return render(request, 'add_user.html')  # Ensure form is empty when not submitting

def user_logout(request):
    # Clear the session variable so that the modal shows again after logging in
    if 'modal_shown' in request.session:
        del request.session['modal_shown']
    logout(request)
    return HttpResponseRedirect('/login')  # Redirect to login page or home page

@login_required
def out_of_stock(request):
    out_of_stock_products = Drug.objects.filter(stock=0)
    return render(request, 'Inventory/out_of_stock.html', {'out_of_stock': out_of_stock_products})

@login_required
def expiring_soon(request):
    today = timezone.now().date()
    expiring_products = Drug.objects.filter(expiry_date__lte=today + timedelta(days=180), stock__gt=0).order_by('expiry_date')
    return render(request, 'Inventory/expiring_soon.html', {'expiring_soon': expiring_products})

blue_shades = [
    {"hex": "#0047AB", "rgba": "rgba(0, 71, 171, 1)"},
    {"hex": "#007b83", "rgba": "rgba(0, 123, 131, 1)"},
    {"hex": "#0275d8", "rgba": "rgba(2, 117, 216, 1)"},
    {"hex": "#4682b4", "rgba": "rgba(70, 130, 180, 1)"},
    {"hex": "#5f9ea0", "rgba": "rgba(95, 158, 160, 1)"},
    {"hex": "#006994", "rgba": "rgba(0, 105, 148, 1)"},
]

other_colors = [
    {"hex": "#fbe7e4", "rgba": "rgba(251, 231, 228, 1)"},
    {"hex": "#e8f8f5", "rgba": "rgba(232, 248, 245, 1)"},
    {"hex": "#fff5e6", "rgba": "rgba(255, 245, 230, 1)"},
    {"hex": "#f9e6ff", "rgba": "rgba(249, 230, 255, 1)"},
    {"hex": "#fff9e6", "rgba": "rgba(255, 249, 230, 1)"},
    {"hex": "#ffebf0", "rgba": "rgba(255, 235, 240, 1)"},
]

def show_colors(request):
    context = {
        'blue_colors': blue_shades,
        'other_colors': other_colors
    }
    return render(request, 'Inventory/colors.html', context)

@login_required
def bin_filter(request):
    """
    Filters sales data based on the date range provided by the user.
    """
    if request.method == 'POST':
        start_date = request.POST.get('start_date', None)
        end_date = request.POST.get('end_date', None)
        
        # Convert string dates to Python date objects
        if start_date:
            start_date = parse_date(start_date)
        if end_date:
            end_date = parse_date(end_date)

        # Filter the sales by date range
        if start_date and end_date:
            sales = Sale.objects.filter(date_sold__range=[start_date, end_date]).order_by('date_sold')
        elif start_date:
            sales = Sale.objects.filter(date_sold__gte=start_date).order_by('date_sold')
        elif end_date:
            sales = Sale.objects.filter(date_sold__lte=end_date).order_by('date_sold')
        else:
            sales = Sale.objects.all().order_by('date_sold')  # Default to all if no dates provided

    else:
        sales = Sale.objects.all().order_by('date_sold')  # Default to all sales if not a POST request

    return render(request, 'Inventory/bin.html', {'sales': sales})

@csrf_exempt  # Temporarily disable CSRF for this AJAX endpoint
def logout_due_to_inactivity(request):
    if request.method == "POST":
        logout(request)  # Logs the user out
        return JsonResponse({"message": "User logged out due to inactivity"})
    return JsonResponse({"message": "Invalid request"}, status=400)

@login_required
def marketing_items(request):
    """Display the list of marketing items."""
    marketing_items = MarketingItem.objects.all()
    context = {
        'marketing_items': marketing_items
    }
    return render(request, 'Inventory/marketing_items.html', context)

@login_required
def marketing_search(request):
    if request.method == 'POST':
        search_query = request.POST.get('search', '').strip()  # Get the search query from the form
        marketing_items = MarketingItem.objects.filter(name__icontains=search_query)  # Perform a case-insensitive search

        # Pass the search results and query back to the template
        return render(request, 'Inventory/marketing_items.html', {
            'marketing_items': marketing_items,
            'search_query': search_query,
        })
    else:
        return render(request, 'Inventory/marketing_items.html', {
            'marketing_items': [],
            'search_query': '',
        })

@login_required
def issue_item(request):
    if request.method == "POST":
        item_id = request.POST.get("item_id")
        issued_to = request.POST.get("issued_to")
        quantity_issued = request.POST.get("quantity_issued")

        # Validate and process the input
        try:
            quantity_issued = int(quantity_issued)
            marketing_item = get_object_or_404(MarketingItem, id=item_id)

            if quantity_issued > marketing_item.stock:
                messages.error(request, f"Cannot issue more than the available stock for {marketing_item.name}.")
            elif quantity_issued <= 0:
                messages.error(request, f"Invalid quantity issued for {marketing_item.name}.")
            else:
                # Deduct the stock from MarketingItem
                marketing_item.stock -= quantity_issued
                marketing_item.save()

                # Create a new IssuedItem entry
                IssuedItem.objects.create(
                    item=marketing_item.name,
                    stock=marketing_item.stock,
                    issued_to=issued_to,
                    quantity_issued=quantity_issued,
                    issued_by=request.user,
                )

                messages.success(request, f"Issued {quantity_issued} of {marketing_item.name} to {issued_to}.")
        except ValueError:
            messages.error(request, "Invalid quantity issued. Please enter a valid number.")
        except Exception as e:
            messages.error(request, f"An error occurred: {str(e)}")

    # Redirect back to the marketing items page
    return redirect("marketing_items")

def issued_items_report(request):
    """
    View to display all issued items with pagination.
    """
    # Fetch all issued items ordered by date
    issued_items = IssuedItem.objects.all().order_by('-date_issued')
    
    # Pagination (default 10 items per page)
    paginator = Paginator(issued_items, 10)  # Change '10' to your desired items per page
    page_number = request.GET.get('page', 1)
    per_page = request.GET.get('per_page', 10)
    paginator.per_page = int(per_page)

    try:
        issued_items_page = paginator.page(page_number)
    except Exception as e:
        issued_items_page = paginator.page(1)

    context = {
        'issued_items': issued_items_page,
    }
    return render(request, 'Inventory/issued_items_report.html', context)

def issued_items_search(request):
    """
    View to search issued items by query.
    """
    if request.method == 'POST':
        query = request.POST.get('query', '').strip()
        if query:
            # Search in item, issued_to, or issued_by fields
            issued_items = IssuedItem.objects.filter(
                Q(item__icontains=query) |
                Q(issued_to__icontains=query) |
                Q(issued_by__username__icontains=query)
            ).order_by('-date_issued')
        else:
            issued_items = IssuedItem.objects.all()

        # Pagination (default 10 items per page)
        paginator = Paginator(issued_items, 10)
        page_number = request.GET.get('page', 1)
        per_page = request.GET.get('per_page', 10)
        paginator.per_page = int(per_page)

        try:
            issued_items_page = paginator.page(page_number)
        except Exception as e:
            issued_items_page = paginator.page(1)

        context = {
            'issued_items': issued_items_page,
            'query': query,  # Pass the query back to the template
        }
        return render(request, 'Inventory/issued_items_report.html', context)

    return render(request, 'Inventory/issued_items_report.html', {'issued_items': []})

def issued_items_filter(request):
    """
    View to filter issued items by a date range.
    """
    if request.method == 'POST':
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')

        # If both dates are provided, filter by range
        if start_date and end_date:
            try:
                start_date_obj = datetime.strptime(start_date, '%Y-%m-%d')
                end_date_obj = datetime.strptime(end_date, '%Y-%m-%d')
                issued_items = IssuedItem.objects.filter(
                    date_issued__range=(start_date_obj, end_date_obj)
                ).order_by('-date_issued')
            except ValueError:
                issued_items = IssuedItem.objects.all().order_by('-date_issued')
        else:
            # If no valid date range is provided, show all items
            issued_items = IssuedItem.objects.all().order_by('-date_issued')

        # Pagination (default 10 items per page)
        paginator = Paginator(issued_items, 10)
        page_number = request.GET.get('page', 1)
        per_page = request.GET.get('per_page', 10)
        paginator.per_page = int(per_page)

        try:
            issued_items_page = paginator.page(page_number)
        except Exception as e:
            issued_items_page = paginator.page(1)

        context = {
            'issued_items': issued_items_page,
            'start_date': start_date,
            'end_date': end_date,
        }
        return render(request, 'Inventory/issued_items_report.html', context)

    return render(request, 'Inventory/issued_items_report.html', {'issued_items': []})

def create_marketing_item(request):
    if request.method == "POST":
        # Get the data from the form
        name = request.POST.get("name")
        stock = request.POST.get("stock")
        
        # Create a new marketing item
        MarketingItem.objects.create(name=name, stock=stock)
        messages.success(request, f'Marketing item "{name}" has been created successfully.')
        
        # Redirect back to the marketing items page
        return redirect('marketing_items')
    
    # Render the creation form
    return render(request, 'Inventory/create_marketing_item.html')

def picking_list_view(request):
    picking_list = PickingList.objects.all().order_by('-date')
    
    # Filtering by search query
    query = request.GET.get('search', '')
    if query:
        picking_list = picking_list.filter(
            Q(client__icontains=query) |  
            Q(product__icontains=query) |
            Q(batch_no__icontains=query) |
            Q(quantity__icontains=query) |
            Q(date__icontains=query) 
        ).order_by('-date')
    
    # Filtering by date range
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    if start_date and end_date:
        picking_list = picking_list.filter(date__range=[start_date, end_date]).order_by('-date')

    
    # Pagination
    per_page = int(request.GET.get('per_page', 10))
    paginator = Paginator(picking_list, per_page)
    page_number = request.GET.get('page',1)
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'Inventory/picking_list.html', {'picking_list': page_obj})



def add_to_picking_list(request, drug_id):
    if request.method == "POST":
        drug = get_object_or_404(Drug, id=drug_id)
        client = request.POST.get("client", "").strip()
        quantity = request.POST.get("quantity", "0").strip()

        # Validate client field (ensure it's not empty)
        if not client:
            messages.error(request, "Client name cannot be empty.")
            return redirect("home")

        # Ensure quantity is a valid integer
        if not quantity.isdigit():
            messages.error(request, "Invalid quantity. Please enter a valid number.")
            return redirect("home")

        quantity = int(quantity)

        if quantity <= 0:
            messages.error(request, "Quantity must be greater than zero.")
            return redirect("home")

        # Check if enough stock is available
        if drug.stock < quantity:
            messages.error(request, "Not enough stock available.")
            return redirect("home")

        # Add item to the picking list
        PickingList.objects.create(
            date=timezone.now(),
            client=client,
            product=drug.name,
            batch_no=drug.batch_no,
            quantity=quantity,
        )

        messages.success(request, "Item added to the picking list.")
        return redirect("home")

    return HttpResponse("Invalid request", status=400)


def cannister_list(request):
    cannisters = Cannister.objects.all()
    return render(request, 'Inventory/cannister.html', {'cannisters': cannisters})

@login_required
def issue_cannister(request, cannister_id):
    cannister = get_object_or_404(Cannister, id=cannister_id)
    
    if request.method == "POST":
        client = request.POST.get("client")
        quantity = int(request.POST.get("quantity"))

        if quantity > 0 and quantity <= cannister.stock:
            # Deduct stock
            cannister.stock -= quantity
            cannister.save()

            # Save issuance record
            IssuedCannister.objects.create(
                name=cannister.name,
                batch_no=cannister.batch_no,
                staff_on_duty=request.user,
                client=client,
                quantity=quantity,
                balance=cannister.stock
            )
    
    return redirect('cannister_list')


@login_required
def bin_card(request):
    issued_cannisters = IssuedCannister.objects.all().order_by('-date_issued')

    # Pagination
    per_page = request.GET.get('per_page', 10)
    paginator = Paginator(issued_cannisters, per_page)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'Inventory/cannister_bin.html', {'issued_cannisters': page_obj})

@login_required
def bin_search(request):
    query = request.GET.get('search', '')
    issued_cannisters = IssuedCannister.objects.filter(
        Q(name__icontains=query) | 
        Q(batch_no__icontains=query) |
        Q(client__icontains=query) |
        Q(staff_on_duty__username__icontains=query)
    ).order_by('-date_issued')

    # Pagination
    per_page = request.GET.get('per_page', 10)
    paginator = Paginator(issued_cannisters, per_page)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'Inventory/cannister_bin.html', {'issued_cannisters': page_obj})

@login_required
def can_filter(request):
    if request.method == "POST":
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')

        issued_cannisters = IssuedCannister.objects.all()
        if start_date and end_date:
            issued_cannisters = issued_cannisters.filter(date_issued__range=[start_date, end_date])

        # Pagination
        per_page = request.GET.get('per_page', 10)
        paginator = Paginator(issued_cannisters, per_page)
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)

        return render(request, 'Inventory/cannister_bin.html', {'issued_cannisters': page_obj})
    
    return redirect('bin_card')

@login_required
def return_cannister(request, issued_cannister_id):
    issued_cannister = get_object_or_404(IssuedCannister, id=issued_cannister_id)
    
    if not issued_cannister.action:  # Ensure it's not already returned
        issued_cannister.action = True
        issued_cannister.returned_by = request.user
        issued_cannister.date_returned = timezone.now()
        issued_cannister.save()

        # Restore stock in the cannister model
        cannister = Cannister.objects.get(batch_no=issued_cannister.batch_no)
        cannister.stock += issued_cannister.quantity
        cannister.save()

    return redirect('bin_card')

def search_cannister(request):
    query = request.POST.get('q', '')  # Get search input
    results = []

    if query:
        results = Cannister.objects.filter(
            Q(name__icontains=query) | 
            Q(batch_no__icontains=query) | 
            Q(stock__icontains=query) |   # Search by stock
            Q(litres__icontains=query)    # Search by litres
        )

    return render(request, 'Inventory/cannister.html', {'cannisters': results, 'query': query})
@login_required
def download_top_sold(request):
    # Aggregate total quantity sold for each product
    top_sold_products = (
        Sale.objects.values('drug_sold')
        .annotate(total_quantity=Sum('quantity'))
        .order_by('-total_quantity')
    )

    # Create the HttpResponse object with CSV header
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="top_sold_products.csv"'

    writer = csv.writer(response)
    writer.writerow(['Product Name', 'Total Quantity Sold'])

    for product in top_sold_products:
        writer.writerow([product['drug_sold'], product['total_quantity']])

    return response


# Client Management Views
@login_required
def client_list(request):
    """Display all clients with pagination"""
    per_page = request.GET.get('per_page', 10)
    page_number = request.GET.get('page', 1)
    
    clients = Client.objects.all().order_by('name')
    paginator = Paginator(clients, per_page)
    page_obj = paginator.get_page(page_number)
    
    context = {
        'clients': page_obj,
        'title': 'Client Management'
    }
    return render(request, 'Inventory/client_list.html', context)


@login_required
def create_client(request):
    """Create a new client"""
    if request.method == 'POST':
        try:
            name = request.POST.get('name', '').strip()
            email = request.POST.get('email', '').strip()
            phone = request.POST.get('phone', '').strip()
            
            if not name:
                messages.error(request, 'Client name is required')
                return redirect('create_client')
            
            # Check if client already exists
            if Client.objects.filter(name__iexact=name).exists():
                messages.warning(request, f'Client "{name}" already exists')
                return redirect('client_list')
            
            # Create new client
            client = Client.objects.create(
                name=name,
                email=email if email else None,
                phone=phone if phone else None
            )
            messages.success(request, f'Client "{client.name}" has been successfully created')
            return redirect('client_list')
        except Exception as e:
            messages.error(request, f'An error occurred: {str(e)}')
            return redirect('create_client')
    
    context = {'title': 'Add New Client'}
    return render(request, 'Inventory/create_client.html', context)


@login_required
def edit_client(request, pk):
    """Edit an existing client"""
    client = get_object_or_404(Client, pk=pk)
    
    if request.method == 'POST':
        try:
            name = request.POST.get('name', '').strip()
            email = request.POST.get('email', '').strip()
            phone = request.POST.get('phone', '').strip()
            
            if not name:
                messages.error(request, 'Client name is required')
                return redirect('edit_client', pk=pk)
            
            # Check if another client has this name
            if Client.objects.filter(name__iexact=name).exclude(pk=pk).exists():
                messages.warning(request, f'Another client with name "{name}" already exists')
                return redirect('edit_client', pk=pk)
            
            # Update client
            client.name = name
            client.email = email if email else None
            client.phone = phone if phone else None
            client.save()
            messages.success(request, f'Client "{client.name}" has been successfully updated')
            return redirect('client_list')
        except Exception as e:
            messages.error(request, f'An error occurred: {str(e)}')
            return redirect('edit_client', pk=pk)
    
    context = {
        'client': client,
        'title': 'Edit Client'
    }
    return render(request, 'Inventory/edit_client.html', context)


@login_required
def delete_client(request, pk):
    """Delete a client"""
    client = get_object_or_404(Client, pk=pk)
    
    if request.method == 'POST':
        try:
            client_name = client.name
            client.delete()
            messages.success(request, f'Client "{client_name}" has been successfully deleted')
            return redirect('client_list')
        except Exception as e:
            messages.error(request, f'Cannot delete this client: {str(e)}')
            return redirect('client_list')
    
    context = {
        'client': client,
        'title': 'Delete Client'
    }
    return render(request, 'Inventory/delete_client.html', context)

from openpyxl import load_workbook

template_path = r'c:\Users\Neko\Desktop\pharmTZ\pharmsaver\Glua\Inventory transfer record template11.xlsx'
wb = load_workbook(template_path)

print("Sheets in template:")
for sheet in wb.sheetnames:
    print(f"  - {sheet}")

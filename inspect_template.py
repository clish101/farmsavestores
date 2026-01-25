from openpyxl import load_workbook
import os

template_path = r'c:\Users\Neko\Desktop\pharmTZ\pharmsaver\Glua\Inventory transfer record template11.xlsx'
wb = load_workbook(template_path)
ws = wb.active

print(f"Sheet name: {ws.title}")
print(f"Dimensions: {ws.dimensions}")
print("\nFirst 20 rows of content:")
for row in range(1, 21):
    row_data = []
    for col in range(1, 6):
        cell = ws.cell(row, col)
        row_data.append(f"Col{col}:{cell.value}")
    print(f"Row {row}: {' | '.join(row_data)}")
